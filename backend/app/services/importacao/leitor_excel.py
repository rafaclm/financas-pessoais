"""
Service base para leitura de planilhas Excel.
Abre o arquivo .xlsx em memoria e fornece acesso as abas.
"""
import io
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from loguru import logger


def carregar_workbook(dados_binarios: bytes) -> Workbook:
    """Carrega o arquivo Excel a partir dos bytes recebidos no upload."""
    try:
        return load_workbook(
            filename=io.BytesIO(dados_binarios),
            read_only=True,
            data_only=True  # le valores calculados das formulas
        )
    except Exception as e:
        logger.error(f"Falha ao abrir Excel: {e}")
        raise ValueError(f"Arquivo Excel invalido: {e}")


def listar_abas(workbook: Workbook) -> list:
    """Retorna a lista de nomes das abas."""
    return list(workbook.sheetnames)


def encontrar_aba(workbook: Workbook, nome: str, fuzzy: bool = True) -> str | None:
    """
    Encontra uma aba pelo nome. Se fuzzy=True, busca aproximada
    (ex: 'aportes' encontra 'Aportes', 'APORTES', 'Aportes ').
    """
    if nome in workbook.sheetnames:
        return nome
    if not fuzzy:
        return None
    nome_lower = nome.lower().strip()
    for aba in workbook.sheetnames:
        if aba.lower().strip() == nome_lower:
            return aba
    return None


def ler_aba_como_lista(workbook: Workbook, nome_aba: str, max_linhas: int = 50000) -> list[list]:
    """
    Le uma aba completa como lista de listas (cada linha = lista de celulas).
    Retorna celulas vazias como None.
    """
    aba = workbook[nome_aba]
    linhas = []
    for i, row in enumerate(aba.iter_rows(values_only=True)):
        if i >= max_linhas:
            break
        linhas.append(list(row))
    return linhas


def info_basica_aba(workbook: Workbook, nome_aba: str) -> dict:
    """Retorna informacoes basicas sobre uma aba (linhas, colunas, primeira linha)."""
    aba = workbook[nome_aba]
    return {
        "nome": nome_aba,
        "max_row": aba.max_row,
        "max_column": aba.max_column,
    }


def analisar_planilha(dados_binarios: bytes) -> dict:
    """
    Analisa a estrutura geral do arquivo recebido.
    Retorna lista de abas + indicacao do que e cada uma.
    """
    wb = carregar_workbook(dados_binarios)

    abas_info = []
    for nome in wb.sheetnames:
        info = info_basica_aba(wb, nome)
        # Classifica a aba pelo nome
        nome_lower = nome.lower().strip()
        tipo = "desconhecida"
        if nome_lower in ("2024", "2025", "2026", "2023"):
            tipo = "movimentos_anuais"
        elif "cripto" in nome_lower:
            tipo = "cripto"
        elif "ativos br" in nome_lower or "ativos bra" in nome_lower:
            tipo = "ativos_br"
        elif "ativos eua" in nome_lower or "ativos usa" in nome_lower:
            tipo = "ativos_eua"
        elif "consolida" in nome_lower:
            tipo = "consolidacao"
        elif "aporte" in nome_lower:
            tipo = "aportes"
        elif "provento" in nome_lower:
            tipo = "proventos"

        info["tipo_detectado"] = tipo
        abas_info.append(info)

    return {
        "total_abas": len(wb.sheetnames),
        "abas": abas_info,
    }